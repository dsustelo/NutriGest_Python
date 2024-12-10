import os
import pandas as pd
import openpyxl
import tkinter as tk
from abc import ABC, abstractmethod


# Classe das linhas dos lista_ingredientes da classe Pratos
class LinhasIngredientes:
    def __init__(self, label_ingrediente, entry_ingrediente, var_ingrediente, label_quantidade, entry_quantidade,
                 var_quantidade):
        self.label_ingrediente = label_ingrediente
        self.entry_ingrediente = entry_ingrediente
        self.var_ingrediente = var_ingrediente
        self.label_quantidade = label_quantidade
        self.entry_quantidade = entry_quantidade
        self.var_quantidade = var_quantidade


# Classe para carregar a base de dados e ativá-la
class CarregaDatabase(ABC):
    def loaddatabase(self):
        pass


class CarregarTabelas(CarregaDatabase):
    def loaddb(self):
        dbexcel = openpyxl.load_workbook('tabelanutri.xlsx')
        sheet = dbexcel.active
        return sheet


class CarregarTemplate(CarregaDatabase):
    def __init__(self, prato):
        self.wb = openpyxl.load_workbook('printtemplate.xlsx')
        # Inicio do código do e-fólio global, pergunta 2c (1/2)
        exportaCSV = ExportarCSV.exportar_prato_para_csv(prato)
        # Fim do código do e-fólio global, pergunta 2c (1/2)

    def loaddb(self):
        sheet = self.wb.active
        return sheet

    def closedb(self):
        self.wb.close()

    def save(self):
        self.wb.save('printtemplatetemp.xlsx')


# Classe da base de dados
class Database:
    def __init__(self):
        self.databasematrl = {}

    def loaddatabase(self):
        sheet = CarregarTabelas().loaddb()
        for row in sheet.iter_rows(min_row=4, values_only=True):
            (codigo, produto, energia, energiakj, monodissacaridos, agua, proteina, lipidos, hc, totalhc, acucar,
             acidosorganicos, alcool, amido,
             oligossacaridos, fibra, sfa, mufa, pufa, trans, acidolinoleico, colesterol, retinol, vit_a_total, caroteno,
             vitd,
             a_tocoferol, tiamina, riboflavina, equivalentesdeniacina, niacina, triptofano, vitb6, vitb12, vitc,
             folatos, cinza, na, k, ca, p, mg, fe, zn) = row
            self.databasematrl[produto] = Ingrediente(produto, energia, proteina, hc, acucar, lipidos, sfa, mufa,
                                                      pufa, fibra, na)

    def set(self, valor):
        self.databasematrl = valor

    def get(self):
        return self.databasematrl


# Classe dos ingredientes // Criar setter e getter
class Ingrediente:
    def __init__(self, nome, energia, proteina, hc, acucar, lipidos, sfa,
                 mufa, pufa, fibra, sodio, quantidade=0.0):
        self.nome = nome
        self.energia = energia
        self.proteinas = proteina
        self.hc = hc
        self.acucar = acucar
        self.lipidos = lipidos
        self.sfa = sfa
        self.mufa = mufa
        self.pufa = pufa
        self.fibra = fibra
        self.sodio = sodio
        self.quantidade = quantidade

    def copy(self):
        return Ingrediente(self.nome, self.energia, self.proteinas, self.hc, self.acucar,
                           self.lipidos, self.sfa, self.mufa, self.pufa, self.fibra, self.sodio, self.quantidade)


# Classe dos pratos // criar setter e getter
class Prato:
    def __init__(self, nome, db):
        self.nome = nome
        self.lista_ingredientes = {}
        self.db = db

    def atualiza_prato(self, event, linhaingredientes, arrayvarnutrientes):
        total_nutrientes = {'energia': 0.0, 'proteinas': 0.0, 'hc': 0.0, 'acucar': 0.0, 'lipidos': 0.0,
                            'sfa': 0.0, 'mufa': 0.0, 'pufa': 0.0, 'fibra': 0.0, 'sodio': 0.0}
        if linhaingredientes.entry_quantidade.get() == '':
            linhaingredientes.var_quantidade.set(0.0)
        if linhaingredientes.entry_ingrediente.get() in self.db:
            self.lista_ingredientes[linhaingredientes.entry_ingrediente] = self.db[
                linhaingredientes.var_ingrediente.get()].copy()
            self.lista_ingredientes[
                linhaingredientes.entry_ingrediente].quantidade = linhaingredientes.var_quantidade.get()
        else:
            if linhaingredientes.entry_ingrediente in self.lista_ingredientes:
                self.lista_ingredientes.pop(linhaingredientes.entry_ingrediente)

        for ingrediente in self.lista_ingredientes:
            total_nutrientes['energia'] += round((self.lista_ingredientes[ingrediente].quantidade *
                                                  self.lista_ingredientes[ingrediente].energia) / 100, 2)
            total_nutrientes['proteinas'] += round((self.lista_ingredientes[ingrediente].quantidade *
                                                    self.lista_ingredientes[ingrediente].proteinas) / 100, 2)
            total_nutrientes['hc'] += round((self.lista_ingredientes[ingrediente].quantidade *
                                             self.lista_ingredientes[ingrediente].hc) / 100, 2)
            total_nutrientes['acucar'] += round((self.lista_ingredientes[ingrediente].quantidade *
                                                 self.lista_ingredientes[ingrediente].acucar) / 100, 2)
            total_nutrientes['lipidos'] += round((self.lista_ingredientes[ingrediente].quantidade *
                                                  self.lista_ingredientes[ingrediente].lipidos) / 100, 2)
            total_nutrientes['sfa'] += round((self.lista_ingredientes[ingrediente].quantidade *
                                              self.lista_ingredientes[ingrediente].sfa) / 100, 2)
            total_nutrientes['mufa'] += round((self.lista_ingredientes[ingrediente].quantidade *
                                               self.lista_ingredientes[ingrediente].mufa) / 100, 2)
            total_nutrientes['pufa'] += round((self.lista_ingredientes[ingrediente].quantidade *
                                               self.lista_ingredientes[ingrediente].pufa) / 100, 2)
            total_nutrientes['fibra'] += round((self.lista_ingredientes[ingrediente].quantidade *
                                                self.lista_ingredientes[ingrediente].fibra) / 100, 2)
            total_nutrientes['sodio'] += round((self.lista_ingredientes[ingrediente].quantidade *
                                                self.lista_ingredientes[ingrediente].sodio) / 100, 2)

        arrayvarnutrientes.var_energia.set(total_nutrientes['energia'])
        arrayvarnutrientes.var_proteinas.set(total_nutrientes['proteinas'])
        arrayvarnutrientes.var_hidratos.set(total_nutrientes['hc'])
        arrayvarnutrientes.var_acucar.set(total_nutrientes['acucar'])
        arrayvarnutrientes.var_lipidos_total.set(total_nutrientes['lipidos'])
        arrayvarnutrientes.var_lipidos_saturados.set(total_nutrientes['sfa'])
        arrayvarnutrientes.var_lipidos_mono.set(total_nutrientes['mufa'])
        arrayvarnutrientes.var_lipidos_poli.set(total_nutrientes['pufa'])
        arrayvarnutrientes.var_fibra.set(total_nutrientes['fibra'])
        arrayvarnutrientes.var_sodio.set(total_nutrientes['sodio'])


# Classe da lista de variáveis dos nutrientes do sector do mostrador
class ListaNutrientes:
    def __init__(self, var_energia, var_proteinas, var_hidratos, var_acucar, var_lipidos_total, var_lipidos_saturados,
                 var_lipidos_mono, var_lipidos_poli, var_fibra, var_sodio):
        self.var_energia = var_energia
        self.var_proteinas = var_proteinas
        self.var_hidratos = var_hidratos
        self.var_acucar = var_acucar
        self.var_lipidos_total = var_lipidos_total
        self.var_lipidos_saturados = var_lipidos_saturados
        self.var_lipidos_mono = var_lipidos_mono
        self.var_lipidos_poli = var_lipidos_poli
        self.var_fibra = var_fibra
        self.var_sodio = var_sodio


# Inicio do código do e-fólio global, pergunta 2c (2/2)
class ExportarCSV:
    @staticmethod
    def exportar_prato_para_csv(prato):
        prato_data = {
            'Nome': prato.nome,
            'Ingrediente': [],
            'Quantidade': [],
        }

        for ingrediente, info in prato.lista_ingredientes.items():
            prato_data['Ingrediente'].append(ingrediente.get())
            prato_data['Quantidade'].append(info.quantidade)

        df = pd.DataFrame(prato_data)
        df.to_csv(f'{prato.nome}_export.csv', index=False)
        print(f'Exportado {prato.nome} para {prato.nome}_export.csv')


# Fim do código do e-fólio global, pergunta 2c (2/2)

# Função que imprime o prato no template para a impressora fisica, ainda com bug
def imprime_prato(arrayvarnutrientes, prato):
    # criar ficheiro temporario de um ficheiro excel existente

    wb = CarregarTemplate(prato)
    sheet = wb.loaddb()

    sheet['E28'] = arrayvarnutrientes.var_energia.get()
    sheet['F28'] = arrayvarnutrientes.var_proteinas.get()
    sheet['G28'] = arrayvarnutrientes.var_hidratos.get()
    sheet['H28'] = arrayvarnutrientes.var_acucar.get()
    sheet['I28'] = arrayvarnutrientes.var_lipidos_total.get()
    sheet['J28'] = arrayvarnutrientes.var_lipidos_saturados.get()
    sheet['K28'] = arrayvarnutrientes.var_lipidos_mono.get()
    sheet['L28'] = arrayvarnutrientes.var_lipidos_poli.get()
    sheet['M28'] = arrayvarnutrientes.var_fibra.get()
    sheet['N28'] = arrayvarnutrientes.var_sodio.get()

    wb.save()

    wb.closedb()

    os.startfile('printtemplatetemp.xlsx', 'print')


# Função que apresenta o auto-complete nos entrys dos ingredientes
def auto_complete(event, lista_autocomplete, db):
    # Obtém o texto atual no Entry
    texto_atual = event.widget.get().lower()

    # Limpa a lista de sugestões
    lista_autocomplete.delete(0, tk.END)

    # Verifica se há algo para mostrar na lista
    if texto_atual:
        # Preenche a lista de sugestões com palavras do dicionário que começam com o texto atual
        for ingrediente in db.get():
            if ingrediente is not None:
                if ingrediente.lower().startswith(texto_atual):
                    lista_autocomplete.insert(tk.END, ingrediente)
        # Mostra a lista de sugestões
        lista_autocomplete.place(x=event.widget.winfo_x(), y=event.widget.winfo_y() + event.widget.winfo_height())
        lista_autocomplete.lift()
    else:
        # Oculta a lista de sugestões se não há nada para mostrar
        lista_autocomplete.place_forget()


# Função que passa o valor clicado no auto-complete para a entry
def on_select(event, prato, linhaingredientes, lista_autocomplete, arrayvarnutrientes):
    # Atualiza o Entry com o valor selecionado na lista de sugestões
    valor_selecionado = lista_autocomplete.get(lista_autocomplete.curselection())
    linhaingredientes.var_ingrediente.set(valor_selecionado)
    # Oculta a lista de sugestões após a seleção
    lista_autocomplete.place_forget()
    Prato.atualiza_prato(prato, None, linhaingredientes, arrayvarnutrientes)


# Função que permite apenas numeros e pontos no entry de quantidade e apagar o caracter errado
def only_numbers(event, entry_quantidade):
    if event.char not in '0123456789.':
        entry_quantidade.delete(len(entry_quantidade.get()) - 1, tk.END)


def on_key_release(event, prato, lista_autocomplete, db, linhaingredientes, arrayvarnutrientes):
    auto_complete(event, lista_autocomplete, db)
    prato.atualiza_prato(event, linhaingredientes, arrayvarnutrientes)


def on_quantidade_key_release(event, prato, linhaingredientes, arrayvarnutrientes, entry_quantidade):
    prato.atualiza_prato(event, linhaingredientes, arrayvarnutrientes)
    only_numbers(event, entry_quantidade)
